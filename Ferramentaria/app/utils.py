from flask import current_app
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from io import BytesIO

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in current_app.config['ALLOWED_EXTENSIONS']

EXCEL_STYLES = {
    'header': {
        'fill': PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid"),
        'font': Font(color="FFFFFF", bold=True, size=12),
        'border': Border(
            left=Side(style='medium', color='FFFFFF'),
            right=Side(style='medium', color='FFFFFF'),
            top=Side(style='medium', color='FFFFFF'),
            bottom=Side(style='medium', color='FFFFFF')
        ),
        'alignment': Alignment(horizontal='center', vertical='center')
    },
    'cell': {
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
        'alignment': Alignment(horizontal='center', vertical='center')
    },
    'status_colors': {
        'success': PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),
        'warning': PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid"),
        'danger': PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
    }
}

def gerar_excel_historico(dados, filtros=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico de Trocas"

    ws.column_dimensions['A'].width = 10  
    ws.column_dimensions['B'].width = 15  
    ws.column_dimensions['C'].width = 25  
    ws.column_dimensions['D'].width = 20  
    ws.column_dimensions['E'].width = 15  
    ws.column_dimensions['F'].width = 15  
    ws.column_dimensions['G'].width = 15  

    ws.merge_cells('A1:G1')
    ws['A1'] = 'Relatório de Histórico de Trocas'
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = EXCEL_STYLES['header']['alignment']

    row = 2
    headers = ['Posição', 'Código', 'Operador', 'Data', 'Vida Útil', 'Produção', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = EXCEL_STYLES['header']['fill']
        cell.font = EXCEL_STYLES['header']['font']
        cell.border = EXCEL_STYLES['header']['border']
        cell.alignment = EXCEL_STYLES['header']['alignment']

    for item in dados:
        row += 1
        ws.cell(row=row, column=1, value=item['posicao'])
        ws.cell(row=row, column=2, value=item['codigo'])
        ws.cell(row=row, column=3, value=item['operador'])
        ws.cell(row=row, column=4, value=item['data'])
        ws.cell(row=row, column=5, value=item['vida_util'])
        ws.cell(row=row, column=6, value=item['producao_atual'])
        
        percentual = (item['producao_atual'] / item['vida_util']) * 100 if item['vida_util'] > 0 else 0
        status_cell = ws.cell(row=row, column=7)
        
        if percentual < 60:
            status_cell.fill = EXCEL_STYLES['status_colors']['success']
            status_cell.value = "Normal"
        elif percentual < 85:
            status_cell.fill = EXCEL_STYLES['status_colors']['warning']
            status_cell.value = "Atenção"
        else:
            status_cell.fill = EXCEL_STYLES['status_colors']['danger']
            status_cell.value = "Crítico"

        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = EXCEL_STYLES['cell']['border']
            cell.alignment = EXCEL_STYLES['cell']['alignment']
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
